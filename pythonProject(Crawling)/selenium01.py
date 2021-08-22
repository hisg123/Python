from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import time

driver = webdriver.Chrome()
url = 'https://search.naver.com/search.naver?where=news&query=%EC%9D%B8%ED%85%94&sm=tab_opt&sort=1&photo=0&field=0&pd=3&ds=2021.08.01&de=2021.08.22&docid=&related=0&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so%3Add%2Cp%3Afrom20210801to20210822&is_sug_officeid=0'
driver.get(url)

span_list = []
span_temp = []
press_list = []
press_temp = []
title_list = []
title_temp = []
context_list = []
context_temp = []

cnt = 10
while(cnt!=0):
    #span
    span_list = driver.find_elements_by_xpath('//span[contains(text()," 전")]')
    for span in span_list:
        span_temp.append(span.text)

    #press
    press_list = driver.find_elements_by_xpath('//a[@class="info press"]')
    for press in press_list:
        press_temp.append(press.text)

    #title
    title_list = driver.find_elements_by_xpath('//a[@class="news_tit"]')
    for title in title_list:
         title_temp.append(title.text)

    #context
    context_list = driver.find_elements_by_xpath('//a[@class="api_txt_lines dsc_txt_wrap"]')
    for context in context_list:

        #if start string is '=', need exception handling
        if context.text[0] == '=':
            replace_string = ""
            for i in range(1, len(context.text)):
                replace_string += context.text[i]
            context_temp.append(replace_string)

        #ordinary context case
        else: context_temp.append(context.text)

    #Turning the next page by click btn
    driver.find_element_by_xpath('//i[contains(text(),"다음")]').click()
    cnt -= 1

temp = [[] for j in range(len(span_temp))]

for i in range(len(press_temp)):
    temp[i].append(span_temp[i])
    temp[i].append(press_temp[i])
    temp[i].append(title_temp[i])
    temp[i].append(context_temp[i])

#control size of row and column in xlsx
write_wb = Workbook()
write_ws = write_wb.active

write_ws.column_dimensions['A'].width = 10
write_ws.column_dimensions['B'].width = 30
write_ws.column_dimensions['C'].width = 80
write_ws.column_dimensions['D'].width = 200

#decorate xlsx
for i in range(1, 5):
    cell = write_ws.cell(row=1, column=i)
    cell.alignment = Alignment(horizontal='center')

write_ws['A1'] = '시간'
write_ws['B1'] = '언론사'
write_ws['C1'] = '제목'
write_ws['D1'] = '내용'

#add item in xlsx
for i in temp:
    write_ws.append(i)

write_wb.save('C:/Users/USER/PycharmProjects/pythonProject(Crawling)/인텔.xlsx')
