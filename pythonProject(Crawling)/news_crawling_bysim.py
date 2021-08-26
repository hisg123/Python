from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# 'webdriver_manager' 패키지모듈 다운로드 필요
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from typing import Final, List
from urllib import parse
from datetime import datetime
from datetime import timedelta
import time
import pandas as pd
from openpyxl import load_workbook

CONST_WAIT_TIME: int = 8
DATE_START: Final = datetime(2021, 1, 1)
DATE_END: Final = datetime(2021, 1, 2)
EXCEL_FILE: Final = 'RESULT1.xlsx'


def wait_element_ready(_driver: webdriver, xpath: str, wait: int = CONST_WAIT_TIME) -> WebElement:
    WebDriverWait(_driver, wait).until(
        expected_conditions.presence_of_all_elements_located(
            (By.XPATH, xpath)))
    web_element = _driver.find_element_by_xpath(xpath)
    return web_element


def wait_elements_ready(_driver: webdriver, xpath: str, wait: int = CONST_WAIT_TIME) -> List:
    WebDriverWait(_driver, wait).until(
        expected_conditions.presence_of_all_elements_located(
            (By.XPATH, xpath)))
    web_elements = _driver.find_elements_by_xpath(xpath)
    return web_elements


def save(lv_df):
    lv_wb = load_workbook(EXCEL_FILE)
    lv_ws = lv_wb['Sheet1']
    lv_last_row = lv_ws.max_row
    for lv_i, lv_row in lv_df.iterrows():
        lv_ws.cell(lv_last_row + lv_i + 1, 1).value = lv_row['publisher']
        lv_ws.cell(lv_last_row + lv_i + 1, 2).value = lv_row["title"]
        lv_ws.cell(lv_last_row + lv_i + 1, 3).value = lv_row['summary']
        lv_ws.cell(lv_last_row + lv_i + 1, 4).value = lv_row['publish_date']
        lv_ws.cell(lv_last_row + lv_i + 1, 5).value = f'=HYPERLINK("{lv_row["link"]}","링크 접속")'
    lv_wb.save(EXCEL_FILE)
    lv_wb.close()


if __name__ == "__main__":
    cur_date = DATE_END
    search_date = f'{cur_date.year}.{str(cur_date.month).zfill(2)}.{str(cur_date.day).zfill(2)}'

    options = Options()
    # options.add_argument("--proxy-server=socks5://127.0.0.1:9150")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])  # USB 오류 로그 invisible 하게
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36")
    options.add_argument(
        "app-version=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.75 Safari/537.36")
    options.headless = False  # 브라우저 자동화 작업을 화면에 표시하지 않게 됨

    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)  # 최신 크롬 드라이버 설치
    driver.get(
        'https://search.naver.com/search.naver?where=news&query={}&ds={}&de={}&pd=3&start=1'.format(parse.quote('"삼성전자"'), search_date, search_date))

    df = pd.DataFrame(columns=['publisher', 'title', 'desc', 'publish_date', 'link'])

    try:
        while DATE_START <= cur_date:
            before = ''

            count = 0
            while True:
                if before == driver.current_url:
                    break
                publisher_list = wait_elements_ready(driver, "//a[@class='info press']")
                publisher_list = [x.text.replace("언론사 선정", "") for x in publisher_list]
                title_list = wait_elements_ready(driver, "//a[@class='news_tit']")
                title_list = [x.text for x in title_list]
                desc_list = wait_elements_ready(driver, "//div[@class='news_dsc']")
                desc_list = [x.text for x in desc_list]
                link_list = wait_elements_ready(driver, "//a[@class='news_tit']")
                link_list = [x.get_attribute("href") for x in link_list]

                for i in range(len(publisher_list)):
                    df = df.append(
                        {'publisher': publisher_list[i], 'title': title_list[i], 'summary': desc_list[i],
                         'publish_date': search_date, 'link': link_list[i]}, ignore_index=True)

                if count >= 10:
                    break
                count += 1

                before = driver.current_url
                wait_element_ready(driver, "//a[@class='btn_next']").click()

            cur_date = cur_date - timedelta(days=1)
            search_date = f'{cur_date.year}.{str(cur_date.month).zfill(2)}.{str(cur_date.day).zfill(2)}'
            driver.get('https://search.naver.com/search.naver?where=news&query={}&ds={}&de={}&pd=3&start=1'.format(parse.quote('"삼성전자"'), search_date, search_date))
        save(df)

    except Exception as e:
        save(df)
        print(e)
